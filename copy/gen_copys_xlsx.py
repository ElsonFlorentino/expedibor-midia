# -*- coding: utf-8 -*-
"""Gera o xlsx com as copies e textos de arte da Fase 1 (awareness/marca) da Expedibor."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(__file__), "COPYS_FASE1_AWARENESS.xlsx")

# (ID, Pilar, Canal, Formato, Texto do post/legenda, Texto da arte, CTA)
COPYS = [
    ("01", "Marca / Institucional", "Meta + YouTube", "Vídeo 15s + Estático",
     "Há mais de 15 anos a Expedibor fabrica peças de suspensão em borracha para quem não abre mão de qualidade. Buchas, batentes, coxins, kits de amortecedor e tubos: a linha completa, com certificação ISO 9001.",
     "Qualidade e confiança, peça por peça\nSuspensão em borracha · ISO 9001", "Conheça a Expedibor"),
    ("02", "Produto / Aplicação", "Meta", "Estático",
     "Bucha que assenta certo é serviço que não volta. A linha de buchas Expedibor cobre uma ampla gama de aplicações, com a precisão que a sua oficina precisa.",
     "Buchas Expedibor\nEncaixe certo, serviço que não volta", "Veja a linha completa"),
    ("03", "Produto / Aplicação", "Meta + YouTube", "Vídeo 15s",
     "Tudo o que o amortecedor precisa, num kit só. Os Expedikits chegam montados e prontos para aplicar, economizando o tempo da sua equipe.",
     "Expedikits\nKit de amortecedor montado e pronto pra aplicar", "Conheça os Expedikits"),
    ("04", "Produto / Aplicação", "Meta", "Estático",
     "Coxim bom é o que segura o conjunto sem ruído e sem folga. Os coxins Expedibor são feitos para durar, com borracha de qualidade reconhecida.",
     "Coxins Expedibor\nFirmeza sem ruído", "Veja a linha completa"),
    ("05", "Produto / Aplicação", "Meta", "Estático",
     "Linha Expeditubo: tubos de refrigeração com o acabamento e a vedação que o reparo exige.",
     "Expeditubo\nRefrigeração sob controle", "Conheça a linha"),
    ("06", "Qualidade / ISO 9001", "Meta + YouTube", "Vídeo 15s + Estático",
     "ISO 9001 não é selo na parede, é processo. Cada peça Expedibor passa por controle de qualidade do início ao fim. É por isso que o lojista confia e o mecânico recomenda.",
     "ISO 9001\nQualidade no processo, peça por peça", "Conheça a Expedibor"),
    ("07", "Dica técnica", "Meta + YouTube", "Vídeo 15s / Reels",
     "Dica rápida: ao trocar o amortecedor, troque também os componentes de borracha (coxim, batente e bucha). Peça nova com borracha velha volta pra oficina. Com os Expedikits, já vem tudo junto.",
     "Trocou o amortecedor?\nTroque a borracha também", "Dica Expedibor"),
    ("08", "Bastidores / Fábrica", "Meta + YouTube", "Vídeo 15s + Estático",
     "Por dentro da Expedibor: mais de 10 mil peças por dia saem da nossa fábrica em Ferraz de Vasconcelos. Tecnologia, controle e gente que entende de borracha.",
     "+10 mil peças por dia\nIndústria nacional de suspensão", "Conheça a fábrica"),
    ("09", "Prova / Distribuição", "Meta", "Estático",
     "Onde tem reposição de qualidade, tem Expedibor. Fale com a gente e leve a linha de suspensão em borracha para o seu balcão.",
     "Sua suspensão em borracha começa aqui\nDistribuidores e lojistas", "Seja um parceiro"),
    ("10", "Amplitude de aplicação", "Meta + YouTube", "Vídeo 15s + Estático",
     "Mais de 800 itens em linha. Seja qual for o carro que entra na oficina, tem grande chance de ter peça de suspensão Expedibor para ele.",
     "+800 itens em linha\nCobertura para o seu dia a dia", "Veja a linha completa"),
    ("11", "Bumper 6s (recall)", "YouTube", "Vídeo 6s",
     "Mensagem única e repetível para lembrança de marca.",
     "Expedibor\nSuspensão em borracha · ISO 9001", "(logo de fecho)"),
    ("12", "Marca / Confiança", "Meta", "Estático",
     "Quem aplica sabe a diferença de uma peça que dura. Expedibor: borracha de suspensão feita para o reparo durar.",
     "Feita para o reparo durar\nExpedibor", "Conheça a marca"),
]

wb = Workbook()
ws = wb.active
ws.title = "Fase 1 - Awareness"

# Paleta neutra (NAO usar cores de outras marcas). Acento = azul.
TEAL = "1F4E79"; DARK = "1C1C1C"
headers = ["ID", "Pilar", "Canal", "Formato", "Texto do post (legenda)", "Texto da arte (na imagem)", "CTA"]
widths = [6, 22, 16, 20, 60, 42, 20]

# título
ws.merge_cells("A1:G1")
c = ws["A1"]; c.value = "Expedibor — Copies e Textos de Arte | Fase 1: Awareness / Marca (jun/jul)"
c.font = Font(bold=True, size=13, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=TEAL)
c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height = 26
ws.merge_cells("A2:G2")
c = ws["A2"]; c.value = "Voz B2B, sem travessão, humanizada. Sempre com logo e foto de produto. Trocar nada por dados reais quando os assets chegarem."
c.font = Font(italic=True, size=9, color="555555"); ws.row_dimensions[2].height = 18

# header da tabela
hrow = 3
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=hrow, column=i, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[chr(64+i)].width = widths[i-1]
ws.row_dimensions[hrow].height = 22

# linhas
r = hrow + 1
for row in COPYS:
    for i, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal=("center" if i in (1,3,4) else "left"))
        cell.font = Font(size=10, bold=(i == 1))
        cell.border = border
        if i == 2:
            cell.font = Font(size=10, bold=True, color=TEAL)
    ws.row_dimensions[r].height = 64
    r += 1

ws.freeze_panes = "A4"
wb.save(OUT)
print("salvo:", OUT, "|", len(COPYS), "copies")
